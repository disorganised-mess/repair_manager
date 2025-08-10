#!/usr/bin/env python3
"""
crm_crm_with_dashboard.py

Full CRM application (single-file) with:
- Customers, Equipment, Work Orders, Invoices, Business Info
- Migration system + backups
- QTableView + QSqlTableModel UI (direct editing)
- Export/Import: CSV, JSON, Excel (.xlsx) with preview and upsert
- PDF generation for invoices and work orders (ReportLab optional)
- Dashboard tab: Quick Stats, Charts (matplotlib), Upcoming Deadlines
- Requires: PyQt6. Optional: reportlab, openpyxl, matplotlib
"""

import sys
import os
import sqlite3
import shutil
import csv
import json
from datetime import datetime, timedelta
from typing import List, Callable, Tuple, Dict, Any

# optional libs
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    HAVE_OPENPYXL = True
except Exception:
    HAVE_OPENPYXL = False

try:
    import reportlab
    HAVE_REPORTLAB = True
except Exception:
    HAVE_REPORTLAB = False

try:
    import matplotlib
    matplotlib.use("QtAgg")
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
    HAVE_MPL = True
except Exception:
    HAVE_MPL = False

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLineEdit,
    QLabel, QTabWidget, QMessageBox, QTableView, QStyledItemDelegate,
    QComboBox, QPushButton, QFileDialog, QTextEdit, QInputDialog, QDialog,
    QTableWidget, QTableWidgetItem, QDialogButtonBox, QSizePolicy
)
from PyQt6.QtSql import QSqlDatabase, QSqlTableModel
from PyQt6.QtCore import Qt, QModelIndex, QTimer
from PyQt6.QtGui import QColor, QBrush, QPainter

# ---------- Configuration ----------
DB_FILE = "crm_data.sqlite"
BACKUP_DIR = "backups"
BACKUP_RETENTION_DAYS = 15
# -----------------------------------

# -------------------------
# Backup utilities
# -------------------------
def ensure_backups_folder():
    try:
        os.makedirs(BACKUP_DIR, exist_ok=True)
    except Exception as e:
        print("Could not create backups directory:", e)

def cleanup_old_backups(retention_days: int = BACKUP_RETENTION_DAYS):
    cutoff = datetime.now() - timedelta(days=retention_days)
    try:
        for fname in os.listdir(BACKUP_DIR):
            path = os.path.join(BACKUP_DIR, fname)
            if not os.path.isfile(path):
                continue
            try:
                mtime = datetime.fromtimestamp(os.path.getmtime(path))
                if mtime < cutoff:
                    os.remove(path)
                    print(f"Removed old backup: {path}")
            except Exception as e:
                print("Error while checking/removing backup:", path, e)
    except FileNotFoundError:
        pass
    except Exception as e:
        print("Failed to clean backups:", e)

def timestamped_backup_name(prefix="crm_backup"):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{ts}.sqlite"

def backup_db_before_migration():
    ensure_backups_folder()
    cleanup_old_backups()
    if os.path.exists(DB_FILE):
        dst = os.path.join(BACKUP_DIR, timestamped_backup_name("crm_backup"))
        try:
            shutil.copy2(DB_FILE, dst)
            print(f"Backup created: {dst}")
        except Exception as e:
            print("Backup failed:", e)
    else:
        print("No existing DB to back up.")

def backup_db_before_import():
    ensure_backups_folder()
    dst = os.path.join(BACKUP_DIR, timestamped_backup_name("crm_import_backup"))
    try:
        if os.path.exists(DB_FILE):
            shutil.copy2(DB_FILE, dst)
            print(f"Import-time backup created: {dst}")
        else:
            print("No existing DB to back up (import-time).")
    except Exception as e:
        print("Import backup failed:", e)

# -------------------------
# Migration system (versioned)
# -------------------------
def ensure_schema_version_table():
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS schema_version (
            version INTEGER NOT NULL
        )
    """)
    cur.execute("SELECT COUNT(*) FROM schema_version")
    if cur.fetchone()[0] == 0:
        cur.execute("INSERT INTO schema_version (version) VALUES (0)")
    conn.commit()
    conn.close()

def get_schema_version() -> int:
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute("SELECT version FROM schema_version LIMIT 1")
    row = cur.fetchone()
    conn.close()
    return int(row[0]) if row else 0

def set_schema_version(v: int):
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute("UPDATE schema_version SET version=?", (v,))
    conn.commit()
    conn.close()

def migration_000_initial_schema(conn: sqlite3.Connection):
    cur = conn.cursor()
    cur.execute("""
    CREATE TABLE IF NOT EXISTS customers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        first_name TEXT,
        last_name TEXT,
        phone TEXT,
        email TEXT,
        address TEXT,
        notes TEXT
    )
    """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS equipment (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        customer_id INTEGER,
        serial_number TEXT,
        cpu TEXT,
        ram TEXT,
        storage TEXT,
        os TEXT,
        notes TEXT,
        FOREIGN KEY(customer_id) REFERENCES customers(id)
    )
    """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS work_orders (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        customer_id INTEGER,
        equipment_id INTEGER,
        description TEXT,
        status TEXT,
        date_created TEXT,
        date_completed TEXT,
        due_date TEXT,
        FOREIGN KEY(customer_id) REFERENCES customers(id),
        FOREIGN KEY(equipment_id) REFERENCES equipment(id)
    )
    """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS invoices (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        work_order_id INTEGER,
        amount REAL,
        status TEXT,
        due_date TEXT,
        notes TEXT,
        FOREIGN KEY(work_order_id) REFERENCES work_orders(id)
    )
    """)
    cur.execute("""
    CREATE TABLE IF NOT EXISTS business_info (
        id INTEGER PRIMARY KEY CHECK (id = 1),
        name TEXT,
        address TEXT,
        phone TEXT,
        email TEXT,
        website TEXT
    )
    """)
    cur.execute("SELECT COUNT(*) FROM business_info")
    if cur.fetchone()[0] == 0:
        cur.execute("INSERT INTO business_info (id, name, address, phone, email, website) VALUES (1, 'Business Name Here', 'Address line 1\\nAddress line 2', 'Phone', 'email@example.com', 'https://example.com')")
    conn.commit()
    print("Migration 000: initial schema ensured.")

def _table_columns(conn: sqlite3.Connection, table_name: str) -> List[Tuple]:
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table_name})")
    return cur.fetchall()

def migration_001_add_website_to_businessinfo(conn: sqlite3.Connection):
    cols = [c[1] for c in _table_columns(conn, "business_info")]
    if "website" not in cols:
        cur = conn.cursor()
        cur.execute("ALTER TABLE business_info ADD COLUMN website TEXT")
        conn.commit()
        print("Migration 001: Added 'website' to business_info")
    else:
        print("Migration 001: 'website' already present")

def migration_002_add_due_date_to_invoices(conn: sqlite3.Connection):
    cols = [c[1] for c in _table_columns(conn, "invoices")]
    if "due_date" not in cols:
        cur = conn.cursor()
        cur.execute("ALTER TABLE invoices ADD COLUMN due_date TEXT")
        conn.commit()
        print("Migration 002: Added 'due_date' to invoices")
    else:
        print("Migration 002: 'due_date' already present")

def migration_003_add_notes_to_equipment(conn: sqlite3.Connection):
    cols = [c[1] for c in _table_columns(conn, "equipment")]
    if "notes" not in cols:
        cur = conn.cursor()
        cur.execute("ALTER TABLE equipment ADD COLUMN notes TEXT")
        conn.commit()
        print("Migration 003: Added 'notes' to equipment")
    else:
        print("Migration 003: 'notes' already present")

def migration_004_add_status_dates_to_workorders(conn: sqlite3.Connection):
    cols = [c[1] for c in _table_columns(conn, "work_orders")]
    cur = conn.cursor()
    changed = False
    if "status" not in cols:
        cur.execute("ALTER TABLE work_orders ADD COLUMN status TEXT"); changed = True
    if "date_created" not in cols:
        cur.execute("ALTER TABLE work_orders ADD COLUMN date_created TEXT"); changed = True
    if "date_completed" not in cols:
        cur.execute("ALTER TABLE work_orders ADD COLUMN date_completed TEXT"); changed = True
    if "due_date" not in cols:
        cur.execute("ALTER TABLE work_orders ADD COLUMN due_date TEXT"); changed = True
    if changed:
        conn.commit()
        print("Migration 004: Added status/date/due_date columns to work_orders")
    else:
        print("Migration 004: work_orders already have required columns")

def migration_005_add_amount_and_notes_to_invoices(conn: sqlite3.Connection):
    cols = [c[1] for c in _table_columns(conn, "invoices")]
    cur = conn.cursor()
    changed = False
    if "amount" not in cols:
        cur.execute("ALTER TABLE invoices ADD COLUMN amount REAL"); changed = True
    if "notes" not in cols:
        cur.execute("ALTER TABLE invoices ADD COLUMN notes TEXT"); changed = True
    if "due_date" not in cols:
        cur.execute("ALTER TABLE invoices ADD COLUMN due_date TEXT"); changed = True
    if changed:
        conn.commit()
        print("Migration 005: Added amount/notes/due_date to invoices")
    else:
        print("Migration 005: invoices already have amount/notes/due_date")

MIGRATIONS: List[Callable[[sqlite3.Connection], None]] = [
    migration_000_initial_schema,
    migration_001_add_website_to_businessinfo,
    migration_002_add_due_date_to_invoices,
    migration_003_add_notes_to_equipment,
    migration_004_add_status_dates_to_workorders,
    migration_005_add_amount_and_notes_to_invoices,
]

def run_migrations():
    print("Starting migrations...")
    ensure_backups_folder()
    backup_db_before_migration()
    ensure_schema_version_table()
    current = get_schema_version()
    print(f"Current DB schema version: {current}, latest migration version: {len(MIGRATIONS)}")
    if current < len(MIGRATIONS):
        conn = sqlite3.connect(DB_FILE)
        try:
            for v in range(current, len(MIGRATIONS)):
                print(f"Running migration {v}: {MIGRATIONS[v].__name__}")
                try:
                    MIGRATIONS[v](conn)
                    set_schema_version(v + 1)
                except Exception as e:
                    print(f"Migration {v} failed: {e}")
                    raise
            print("All migrations applied.")
        finally:
            conn.close()
    else:
        print("No migrations needed.")

# -------------------------
# PDF helpers (ReportLab)
# -------------------------
def have_reportlab():
    return HAVE_REPORTLAB

def generate_invoice_pdf(invoice_id: int, save_path: str):
    if not HAVE_REPORTLAB:
        raise RuntimeError("reportlab not installed")
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    conn = sqlite3.connect(DB_FILE); cur = conn.cursor()
    cur.execute("SELECT name, address, phone, email, website FROM business_info WHERE id=1")
    b = cur.fetchone() or ("Business Name", "Address", "Phone", "Email", "")
    b_name, b_addr, b_phone, b_email, b_web = b
    cur.execute("SELECT id, work_order_id, amount, status, due_date, notes FROM invoices WHERE id=?", (invoice_id,))
    inv = cur.fetchone()
    if not inv:
        conn.close()
        raise RuntimeError("Invoice not found")
    inv_id, wo_id, amount, status, due_date, notes = inv
    cust_info = ("", "", "", "", "")
    wo_desc = None
    equip_info = None
    if wo_id:
        cur.execute("SELECT customer_id, equipment_id, description FROM work_orders WHERE id=?", (wo_id,))
        wo = cur.fetchone()
        if wo:
            cust_id, equip_id, wo_desc = wo
            cur.execute("SELECT first_name, last_name, phone, email, address FROM customers WHERE id=?", (cust_id,))
            c = cur.fetchone()
            if c:
                cust_info = (c[0] or "", c[1] or "", c[2] or "", c[3] or "", c[4] or "")
            if equip_id:
                cur.execute("SELECT serial_number, cpu, ram, storage, os FROM equipment WHERE id=?", (equip_id,))
                e = cur.fetchone()
                if e:
                    equip_info = {"serial": e[0], "cpu": e[1], "ram": e[2], "storage": e[3], "os": e[4]}
    conn.close()
    doc = SimpleDocTemplate(save_path, pagesize=letter)
    styles = getSampleStyleSheet()
    flow = []
    flow.append(Paragraph(b_name or "Business Name", styles["Title"]))
    if b_addr:
        for line in str(b_addr).splitlines():
            flow.append(Paragraph(line, styles["Normal"]))
    flow.append(Paragraph(f"Phone: {b_phone or ''}   Email: {b_email or ''}", styles["Normal"]))
    if b_web:
        flow.append(Paragraph(f"Website: {b_web}", styles["Normal"]))
    flow.append(Spacer(1, 12))
    flow.append(Paragraph(f"<b>Invoice #{inv_id}</b>", styles["Heading2"]))
    flow.append(Paragraph(f"Status: {status or ''}", styles["Normal"]))
    flow.append(Paragraph(f"Date: {datetime.now().strftime('%Y-%m-%d')}", styles["Normal"]))
    if due_date:
        flow.append(Paragraph(f"Due date: {due_date}", styles["Normal"]))
    flow.append(Spacer(1, 8))
    flow.append(Paragraph("<b>Customer</b>", styles["Heading3"]))
    flow.append(Paragraph(f"{cust_info[0]} {cust_info[1]}", styles["Normal"]))
    if cust_info[4]:
        flow.append(Paragraph(cust_info[4], styles["Normal"]))
    flow.append(Paragraph(f"Phone: {cust_info[2]}", styles["Normal"]))
    flow.append(Paragraph(f"Email: {cust_info[3]}", styles["Normal"]))
    flow.append(Spacer(1, 8))
    if wo_id:
        flow.append(Paragraph("<b>Related Work Order</b>", styles["Heading3"]))
        flow.append(Paragraph(f"WO #{wo_id}", styles["Normal"]))
        if wo_desc:
            flow.append(Paragraph(f"Description: {wo_desc}", styles["Normal"]))
        flow.append(Spacer(1, 8))
    flow.append(Paragraph("<b>Amount</b>", styles["Heading3"]))
    flow.append(Paragraph(f"${amount if amount is not None else 0:.2f}", styles["Normal"]))
    flow.append(Spacer(1, 8))
    if notes:
        flow.append(Paragraph("<b>Notes</b>", styles["Heading3"]))
        flow.append(Paragraph(notes, styles["Normal"]))
    if equip_info:
        flow.append(Spacer(1, 8))
        flow.append(Paragraph("<b>Equipment</b>", styles["Heading3"]))
        for k, v in equip_info.items():
            flow.append(Paragraph(f"{k.capitalize()}: {v or ''}", styles["Normal"]))
    doc.build(flow)

def generate_workorder_pdf(workorder_id: int, save_path: str):
    if not HAVE_REPORTLAB:
        raise RuntimeError("reportlab not installed")
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    conn = sqlite3.connect(DB_FILE); cur = conn.cursor()
    cur.execute("SELECT name, address, phone, email, website FROM business_info WHERE id=1")
    b = cur.fetchone() or ("Business Name", "Address", "Phone", "Email", "")
    b_name, b_addr, b_phone, b_email, b_web = b
    cur.execute("SELECT id, customer_id, equipment_id, description, status, date_created, date_completed FROM work_orders WHERE id=?", (workorder_id,))
    wo = cur.fetchone()
    if not wo:
        conn.close()
        raise RuntimeError("Work order not found")
    wo_id, cust_id, equip_id, desc, status, date_created, date_completed = wo
    cust_info = ("", "", "", "", "")
    if cust_id:
        cur.execute("SELECT first_name, last_name, phone, email, address FROM customers WHERE id=?", (cust_id,))
        c = cur.fetchone()
        if c:
            cust_info = (c[0] or "", c[1] or "", c[2] or "", c[3] or "", c[4] or "")
    equip_info = None
    if equip_id:
        cur.execute("SELECT serial_number, cpu, ram, storage, os, notes FROM equipment WHERE id=?", (equip_id,))
        e = cur.fetchone()
        if e:
            equip_info = {"serial": e[0], "cpu": e[1], "ram": e[2], "storage": e[3], "os": e[4], "notes": e[5]}
    conn.close()
    doc = SimpleDocTemplate(save_path, pagesize=letter)
    styles = getSampleStyleSheet()
    flow = []
    flow.append(Paragraph(b_name or "Business Name", styles["Title"]))
    if b_addr:
        for line in str(b_addr).splitlines():
            flow.append(Paragraph(line, styles["Normal"]))
    flow.append(Paragraph(f"Phone: {b_phone or ''}   Email: {b_email or ''}", styles["Normal"]))
    if b_web:
        flow.append(Paragraph(f"Website: {b_web}", styles["Normal"]))
    flow.append(Spacer(1, 12))
    flow.append(Paragraph(f"<b>Work Order #{wo_id}</b>", styles["Heading2"]))
    flow.append(Paragraph(f"Status: {status or ''}", styles["Normal"]))
    flow.append(Paragraph(f"Date created: {date_created or ''}", styles["Normal"]))
    if date_completed:
        flow.append(Paragraph(f"Date completed: {date_completed}", styles["Normal"]))
    flow.append(Spacer(1, 8))
    flow.append(Paragraph("<b>Customer</b>", styles["Heading3"]))
    flow.append(Paragraph(f"{cust_info[0]} {cust_info[1]}", styles["Normal"]))
    if cust_info[4]:
        flow.append(Paragraph(cust_info[4], styles["Normal"]))
    flow.append(Paragraph(f"Phone: {cust_info[2]}", styles["Normal"]))
    flow.append(Paragraph(f"Email: {cust_info[3]}", styles["Normal"]))
    flow.append(Spacer(1, 8))
    flow.append(Paragraph("<b>Work Description</b>", styles["Heading3"]))
    flow.append(Paragraph(desc or "", styles["Normal"]))
    flow.append(Spacer(1, 8))
    if equip_info:
        flow.append(Paragraph("<b>Equipment</b>", styles["Heading3"]))
        for k in ("serial", "cpu", "ram", "storage", "os"):
            flow.append(Paragraph(f"{k.capitalize()}: {equip_info.get(k) or ''}", styles["Normal"]))
        if equip_info.get("notes"):
            flow.append(Paragraph("<b>Equipment Notes</b>", styles["Heading3"]))
            flow.append(Paragraph(equip_info.get("notes") or "", styles["Normal"]))
    doc.build(flow)

# -------------------------
# Delegates
# -------------------------
class ComboBoxDelegate(QStyledItemDelegate):
    def __init__(self, items: List[str], parent=None):
        super().__init__(parent)
        self.items = items
    def createEditor(self, parent, option, index):
        combo = QComboBox(parent)
        combo.addItems(self.items)
        combo.currentIndexChanged.connect(lambda *_: self.commitData.emit(combo))
        return combo
    def setEditorData(self, editor: QComboBox, index: QModelIndex):
        value = index.data(Qt.ItemDataRole.DisplayRole)
        if value is None:
            value = ""
        idx = editor.findText(str(value))
        if idx >= 0:
            editor.setCurrentIndex(idx)
    def setModelData(self, editor: QComboBox, model, index: QModelIndex):
        model.setData(index, editor.currentText(), Qt.ItemDataRole.EditRole)
        model.dataChanged.emit(index, index)

class HighlightDelegate(QStyledItemDelegate):
    def __init__(self, view, parent=None):
        super().__init__(parent)
        self.view = view
    def paint(self, painter: QPainter, option, index: QModelIndex):
        text = str(index.data(Qt.ItemDataRole.DisplayRole) or "")
        search_text = getattr(self.view, "search_text", "")
        if search_text and search_text.lower() in text.lower():
            painter.save()
            painter.fillRect(option.rect, QBrush(QColor(255, 255, 150, 140)))
            painter.restore()
        super().paint(painter, option, index)

class StatusColorDelegate(QStyledItemDelegate):
    def __init__(self, mapping: dict, parent=None):
        super().__init__(parent)
        self.mapping = {k.lower(): v for k, v in mapping.items()}
    def initStyleOption(self, option, index: QModelIndex):
        super().initStyleOption(option, index)
        val = index.data(Qt.ItemDataRole.DisplayRole)
        if isinstance(val, str):
            v = val.lower()
            if v in self.mapping:
                option.backgroundBrush = QBrush(self.mapping[v])

# -------------------------
# Export / Import helpers (CSV/JSON/Excel)
# -------------------------
def table_columns_sql(table: str) -> List[str]:
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table})")
    cols = [row[1] for row in cur.fetchall()]
    conn.close()
    return cols

def export_table_csv(table: str, path: str) -> int:
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute(f"SELECT * FROM {table}")
    rows = cur.fetchall()
    cols = [d[0] for d in cur.description]
    conn.close()
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(cols)
        writer.writerows(rows)
    return len(rows)

def export_table_json(table: str, path: str) -> int:
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute(f"SELECT * FROM {table}")
    rows = cur.fetchall()
    cols = [d[0] for d in cur.description]
    conn.close()
    data = [dict(zip(cols, row)) for row in rows]
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return len(rows)

def export_table_excel(table: str, path: str) -> int:
    if not HAVE_OPENPYXL:
        raise RuntimeError("openpyxl not installed")
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    cur.execute(f"SELECT * FROM {table}")
    rows = cur.fetchall()
    cols = [d[0] for d in cur.description]
    conn.close()
    wb = Workbook()
    ws = wb.active
    ws.title = table[:31]
    ws.append(cols)
    for row in rows:
        ws.append(list(row))
    for i, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            val = cell.value
            if val is None:
                l = 0
            else:
                l = len(str(val))
            if l > max_len:
                max_len = l
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(max_len + 2, 100)
    wb.save(path)
    return len(rows)

def detect_file_type_by_ext(path: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return "excel"
    if ext == ".json":
        return "json"
    return "csv"

def read_csv_file(path: str) -> Tuple[List[str], List[List[str]]]:
    with open(path, newline='', encoding='utf-8') as f:
        rdr = csv.DictReader(f, skipinitialspace=True)
        header = rdr.fieldnames or []
        rows = []
        for row in rdr:
            rows.append([row.get(h, "") for h in header])
    return header, rows

def read_json_file(path: str) -> Tuple[List[str], List[List[str]]]:
    with open(path, "r", encoding='utf-8') as f:
        data = json.load(f)
    if not isinstance(data, list) or (len(data) > 0 and not isinstance(data[0], dict)):
        raise ValueError("JSON import expects a list of objects (dictionaries).")
    header = list(data[0].keys())
    rows = []
    for item in data:
        rows.append([item.get(h, "") for h in header])
    return header, rows

def read_excel_file(path: str) -> Tuple[List[str], List[List[str]]]:
    if not HAVE_OPENPYXL:
        raise RuntimeError("openpyxl not installed")
    wb = load_workbook(filename=path, data_only=True)
    ws = wb.active
    rows = list(ws.rows)
    if not rows:
        return [], []
    header = [str(cell.value) if cell.value is not None else "" for cell in rows[0]]
    data_rows = []
    for r in rows[1:]:
        data_rows.append([cell.value if cell.value is not None else "" for cell in r])
    data_rows = [[str(v) if v is not None else "" for v in row] for row in data_rows]
    return header, data_rows

def validate_import_columns(table: str, header: List[str]) -> Tuple[bool, List[str]]:
    cols = table_columns_sql(table)
    missing = [h for h in header if h not in cols]
    return (len(missing) == 0), missing

def coerce_value_by_type(value: str, col_type: str):
    if value is None:
        return None
    s = str(value).strip()
    if s == "":
        return None
    t = (col_type or "").upper()
    if "INT" in t:
        try:
            return int(float(s))
        except Exception:
            return s
    if "REAL" in t or "FLOA" in t or "DOUB" in t:
        try:
            return float(s)
        except Exception:
            return s
    return s

def upsert_rows_into_table(table: str, header: List[str], rows: List[List[str]]) -> Tuple[int, int]:
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()
    pragma = _table_columns(conn, table)
    col_types = {r[1]: r[2] for r in pragma}
    pk_cols = [r[1] for r in pragma if r[5] == 1]
    pk_col = pk_cols[0] if pk_cols else "id"
    inserted = 0
    updated = 0
    conn.execute("BEGIN")
    try:
        for r in rows:
            data = dict(zip(header, r))
            typed = {col: coerce_value_by_type(data.get(col, None), col_types.get(col, "")) for col in header}
            if pk_col in header and typed.get(pk_col) is not None:
                pk_val = typed.get(pk_col)
                cur.execute(f"SELECT 1 FROM {table} WHERE {pk_col} = ?", (pk_val,))
                if cur.fetchone():
                    set_pairs = ", ".join([f"`{col}` = ?" for col in header if col != pk_col])
                    vals = [typed[col] for col in header if col != pk_col]
                    vals.append(pk_val)
                    try:
                        cur.execute(f"UPDATE {table} SET {set_pairs} WHERE {pk_col} = ?", vals)
                        updated += 1
                    except Exception as e:
                        print("Update error:", e, "row:", data)
                        continue
                    continue
            insert_cols = [c for c in header if not (c == pk_col and (typed.get(pk_col) is None))]
            placeholders = ", ".join(["?"] * len(insert_cols))
            cols_sql = ", ".join([f"`{c}`" for c in insert_cols])
            vals = [typed[c] for c in insert_cols]
            try:
                cur.execute(f"INSERT INTO {table} ({cols_sql}) VALUES ({placeholders})", vals)
                inserted += 1
            except Exception as e:
                print("Insert error:", e, "row:", data)
                continue
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        raise
    conn.close()
    return inserted, updated

# -------------------------
# Import Preview Dialog
# -------------------------
class ImportPreviewDialog(QDialog):
    def __init__(self, parent, table: str, header: List[str], rows: List[List[Any]]):
        super().__init__(parent)
        self.setWindowTitle(f"Import Preview — {table}")
        self.table_name = table
        self.header = header
        self.rows = rows
        self.resize(1000, 520)
        v = QVBoxLayout(self)
        info = QLabel(f"{len(rows)} rows, {len(header)} columns. Edit cells if needed before import.")
        v.addWidget(info)
        self.table = QTableWidget()
        self.table.setColumnCount(len(header))
        self.table.setRowCount(len(rows))
        self.table.setHorizontalHeaderLabels(header)
        for r_idx, row in enumerate(rows):
            for c_idx, val in enumerate(row):
                item = QTableWidgetItem("" if val is None else str(val))
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(r_idx, c_idx, item)
        self.table.resizeColumnsToContents()
        self.table.horizontalHeader().setStretchLastSection(True)
        v.addWidget(self.table)
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("Import")
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        v.addWidget(btns)
    def get_preview_rows(self) -> Tuple[List[str], List[List[str]]]:
        header = [self.table.horizontalHeaderItem(c).text() for c in range(self.table.columnCount())]
        rows = []
        for r in range(self.table.rowCount()):
            rowdata = []
            for c in range(self.table.columnCount()):
                it = self.table.item(r, c)
                rowdata.append(it.text() if it else "")
            rows.append(rowdata)
        return header, rows

# -------------------------
# Dashboard Widget
# -------------------------
class DashboardWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.layout = QVBoxLayout(self)
        self.stats_layout = QHBoxLayout()
        self.layout.addLayout(self.stats_layout)
        # create stat labels
        self.stat_labels = {}
        for key in ("total_customers", "total_equipment", "total_work_orders", "pending_work_orders",
                    "total_invoices", "outstanding_count", "outstanding_total", "paid_count", "paid_total"):
            lbl = QLabel("..."); lbl.setWordWrap(True)
            self.stat_labels[key] = lbl
            self.stats_layout.addWidget(lbl)
        # charts area
        self.charts_layout = QHBoxLayout()
        self.layout.addLayout(self.charts_layout)
        if HAVE_MPL:
            self.wo_fig = Figure(figsize=(4,3))
            self.wo_canvas = FigureCanvas(self.wo_fig)
            self.inv_fig = Figure(figsize=(4,3))
            self.inv_canvas = FigureCanvas(self.inv_fig)
            self.charts_layout.addWidget(self.wo_canvas)
            self.charts_layout.addWidget(self.inv_canvas)
        else:
            self.charts_layout.addWidget(QLabel("matplotlib not installed - charts disabled"))
        # upcoming deadlines lists
        self.deadline_layout = QHBoxLayout()
        self.layout.addLayout(self.deadline_layout)
        self.wo_deadlines = QTableWidget()
        self.wo_deadlines.setColumnCount(3)
        self.wo_deadlines.setHorizontalHeaderLabels(["WO ID","Due Date","Description"])
        self.inv_deadlines = QTableWidget()
        self.inv_deadlines.setColumnCount(4)
        self.inv_deadlines.setHorizontalHeaderLabels(["Inv ID","Due Date","Status","Amount"])
        self.deadline_layout.addWidget(self.wo_deadlines)
        self.deadline_layout.addWidget(self.inv_deadlines)
        # refresh button
        btn_layout = QHBoxLayout()
        self.refresh_btn = QPushButton("Refresh")
        self.refresh_btn.clicked.connect(self.refresh)
        btn_layout.addStretch()
        btn_layout.addWidget(self.refresh_btn)
        self.layout.addLayout(btn_layout)
        # initial refresh
        QTimer.singleShot(100, self.refresh)

    def refresh(self):
        # query DB for stats and populate charts + deadlines
        conn = sqlite3.connect(DB_FILE); cur = conn.cursor()
        # stats
        cur.execute("SELECT COUNT(*) FROM customers"); total_customers = cur.fetchone()[0] or 0
        cur.execute("SELECT COUNT(*) FROM equipment"); total_equipment = cur.fetchone()[0] or 0
        cur.execute("SELECT COUNT(*) FROM work_orders"); total_work_orders = cur.fetchone()[0] or 0
        cur.execute("SELECT COUNT(*) FROM work_orders WHERE status LIKE 'Pending' OR status='' OR status IS NULL"); pending_work_orders = cur.fetchone()[0] or 0
        cur.execute("SELECT COUNT(*), COALESCE(SUM(amount),0) FROM invoices"); total_invoices_count, total_invoices_sum = cur.fetchone()
        cur.execute("SELECT COUNT(*), COALESCE(SUM(amount),0) FROM invoices WHERE status LIKE 'Outstanding'"); outstanding_count, outstanding_total = cur.fetchone()
        cur.execute("SELECT COUNT(*), COALESCE(SUM(amount),0) FROM invoices WHERE status LIKE 'Paid'"); paid_count, paid_total = cur.fetchone()
        conn.close()
        self.stat_labels["total_customers"].setText(f"Customers: {total_customers}")
        self.stat_labels["total_equipment"].setText(f"Equipment: {total_equipment}")
        self.stat_labels["total_work_orders"].setText(f"Work Orders: {total_work_orders}")
        self.stat_labels["pending_work_orders"].setText(f"Pending WOs: {pending_work_orders}")
        self.stat_labels["total_invoices"].setText(f"Invoices: {total_invoices_count} (${total_invoices_sum or 0:.2f})")
        self.stat_labels["outstanding_count"].setText(f"Outstanding: {outstanding_count}")
        self.stat_labels["outstanding_total"].setText(f"Outstanding $: ${outstanding_total or 0:.2f}")
        self.stat_labels["paid_count"].setText(f"Paid: {paid_count}")
        self.stat_labels["paid_total"].setText(f"Paid $: ${paid_total or 0:.2f}")
        # charts
        if HAVE_MPL:
            try:
                conn = sqlite3.connect(DB_FILE); cur = conn.cursor()
                cur.execute("SELECT status, COUNT(*) FROM work_orders GROUP BY status")
                rows = cur.fetchall()
                statuses = [r[0] or "Unknown" for r in rows]
                counts = [r[1] for r in rows]
                self.wo_fig.clear()
                ax = self.wo_fig.add_subplot(111)
                ax.bar(statuses, counts)
                ax.set_title("Work Orders by Status")
                ax.set_ylabel("Count")
                self.wo_canvas.draw()
                cur.execute("SELECT status, COUNT(*) FROM invoices GROUP BY status")
                rows = cur.fetchall()
                labels = [r[0] or "Unknown" for r in rows]
                sizes = [r[1] for r in rows]
                self.inv_fig.clear()
                ax2 = self.inv_fig.add_subplot(111)
                if sizes:
                    ax2.pie(sizes, labels=labels, autopct="%1.1f%%")
                ax2.set_title("Invoices by Status")
                self.inv_canvas.draw()
                conn.close()
            except Exception as e:
                print("Chart draw failed:", e)
        # upcoming deadlines
        conn = sqlite3.connect(DB_FILE); cur = conn.cursor()
        today = datetime.now().strftime("%Y-%m-%d")
        cur.execute("SELECT id, due_date, description FROM work_orders WHERE due_date IS NOT NULL AND due_date<>'' ORDER BY due_date ASC LIMIT 5")
        wos = cur.fetchall()
        cur.execute("SELECT id, due_date, status, amount FROM invoices WHERE due_date IS NOT NULL AND due_date<>'' ORDER BY due_date ASC LIMIT 5")
        invs = cur.fetchall()
        conn.close()
        # populate WOs
        self.wo_deadlines.setRowCount(len(wos))
        for r, row in enumerate(wos):
            wid, due, desc = row
            it0 = QTableWidgetItem(str(wid)); it1 = QTableWidgetItem(due or ""); it2 = QTableWidgetItem(desc or "")
            if due and due < today:
                # overdue
                it1.setBackground(QBrush(QColor(255,200,200)))
            self.wo_deadlines.setItem(r, 0, it0); self.wo_deadlines.setItem(r, 1, it1); self.wo_deadlines.setItem(r, 2, it2)
        self.wo_deadlines.resizeColumnsToContents()
        self.wo_deadlines.horizontalHeader().setStretchLastSection(True)
        # populate invoices
        self.inv_deadlines.setRowCount(len(invs))
        for r, row in enumerate(invs):
            iid, due, status, amount = row
            it0 = QTableWidgetItem(str(iid)); it1 = QTableWidgetItem(due or ""); it2 = QTableWidgetItem(status or ""); it3 = QTableWidgetItem(f"${amount or 0:.2f}")
            if due and due < today and (status or "").lower() != "paid":
                it1.setBackground(QBrush(QColor(255,200,200)))
            self.inv_deadlines.setItem(r, 0, it0); self.inv_deadlines.setItem(r, 1, it1); self.inv_deadlines.setItem(r, 2, it2); self.inv_deadlines.setItem(r, 3, it3)
        self.inv_deadlines.resizeColumnsToContents()
        self.inv_deadlines.horizontalHeader().setStretchLastSection(True)

# -------------------------
# Main Application
# -------------------------
class CRMApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("CRM — Full with Dashboard")
        self.resize(1280, 860)
        run_migrations()
        if not QSqlDatabase.isDriverAvailable("QSQLITE"):
            QMessageBox.critical(self, "Error", "SQLite driver not available.")
            sys.exit(1)
        self.db = QSqlDatabase.addDatabase("QSQLITE")
        self.db.setDatabaseName(DB_FILE)
        if not self.db.open():
            QMessageBox.critical(self, "Error", "Could not open database.")
            sys.exit(1)
        central = QWidget(); self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)
        self.tabs = QTabWidget(); main_layout.addWidget(self.tabs)
        # Dashboard first
        self.setup_dashboard_tab()
        self.setup_customers_tab()
        self.setup_equipment_tab()
        self.setup_workorders_tab()
        self.setup_invoices_tab()
        self.setup_businessinfo_tab()
        # refresh dashboard when tab changed to dashboard or after edits (timer)
        self.tabs.currentChanged.connect(self.on_tab_changed)

    def on_tab_changed(self, idx):
        if self.tabs.tabText(idx) == "Dashboard":
            if hasattr(self, "dashboard_widget"):
                self.dashboard_widget.refresh()

    # ---------- Helper UI Export/Import ----------
    def _choose_export_and_save(self, table: str):
        options = ["CSV", "JSON"]
        if HAVE_OPENPYXL:
            options.append("Excel")
        fmt, ok = QInputDialog.getItem(self, "Export format", "Choose format:", options, 0, False)
        if not ok:
            return
        fmt = fmt.lower()
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_ext = "csv" if fmt == "csv" else ("xlsx" if fmt == "excel" else "json")
        default = f"{table}_{ts}.{default_ext}"
        if fmt == "excel":
            filt = "Excel files (*.xlsx)"
        else:
            filt = "CSV files (*.csv);;JSON files (*.json)" if fmt == "csv" else "JSON files (*.json);;CSV files (*.csv)"
        path, _ = QFileDialog.getSaveFileName(self, "Export table", default, filt)
        if not path:
            return
        try:
            if fmt == "csv":
                count = export_table_csv(table, path)
            elif fmt == "json":
                count = export_table_json(table, path)
            else:
                count = export_table_excel(table, path)
            QMessageBox.information(self, "Exported", f"Exported {count} rows from '{table}' to {path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Export failed: {e}")

    def _import_with_preview(self, table: str):
        if HAVE_OPENPYXL:
            filters = "CSV files (*.csv);;JSON files (*.json);;Excel files (*.xlsx)"
        else:
            filters = "CSV files (*.csv);;JSON files (*.json)"
        path, _ = QFileDialog.getOpenFileName(self, "Import table", "", filters)
        if not path:
            return
        ftype = detect_file_type_by_ext(path)
        try:
            if ftype == "csv":
                header, rows = read_csv_file(path)
            elif ftype == "json":
                header, rows = read_json_file(path)
            else:
                header, rows = read_excel_file(path)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to parse file: {e}")
            return
        if not header:
            QMessageBox.warning(self, "Empty", "No columns found in file.")
            return
        ok, missing = validate_import_columns(table, header)
        if not ok:
            QMessageBox.critical(self, "Invalid", f"The file contains columns not present in '{table}': {missing}")
            return
        preview = ImportPreviewDialog(self, table, header, rows)
        if preview.exec() != QDialog.DialogCode.Accepted:
            return
        header2, rows2 = preview.get_preview_rows()
        backup_db_before_import()
        try:
            inserted, updated = upsert_rows_into_table(table, header2, rows2)
            QMessageBox.information(self, "Imported", f"Imported into '{table}': inserted {inserted}, updated {updated}")
            # refresh relevant models/views
            if table == "customers":
                self.customer_model.select(); self.customer_view.resizeColumnsToContents()
            elif table == "equipment":
                self.equipment_model.select(); self.equipment_view.resizeColumnsToContents()
            elif table == "work_orders":
                self.wo_model.select(); self.wo_view.resizeColumnsToContents()
            elif table == "invoices":
                self.inv_model.select(); self.inv_view.resizeColumnsToContents()
            elif table == "business_info":
                self.load_business_info()
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Import failed: {e}")

    # ---------- Dashboard ----------
    def setup_dashboard_tab(self):
        tab = QWidget(); v = QVBoxLayout(tab)
        self.dashboard_widget = DashboardWidget(self)
        v.addWidget(self.dashboard_widget)
        self.tabs.addTab(tab, "Dashboard")

    # ---------- Customers ----------
    def setup_customers_tab(self):
        tab = QWidget(); v = QVBoxLayout(tab)
        top = QHBoxLayout()
        top.addWidget(QLabel("Search:"))
        self.customer_search = QLineEdit(); self.customer_search.setPlaceholderText("search first/last/phone/email/address/notes ...")
        self.customer_search.textChanged.connect(self.on_customer_search)
        top.addWidget(self.customer_search)
        new_btn = QPushButton("New"); del_btn = QPushButton("Delete")
        exp_btn = QPushButton("Export"); imp_btn = QPushButton("Import")
        new_btn.clicked.connect(self.new_customer); del_btn.clicked.connect(self.delete_customer)
        exp_btn.clicked.connect(lambda: self._choose_export_and_save("customers"))
        imp_btn.clicked.connect(lambda: self._import_with_preview("customers"))
        top.addWidget(new_btn); top.addWidget(del_btn); top.addWidget(exp_btn); top.addWidget(imp_btn)
        v.addLayout(top)
        self.customer_model = QSqlTableModel(self, self.db); self.customer_model.setTable("customers")
        self.customer_model.setEditStrategy(QSqlTableModel.EditStrategy.OnFieldChange); self.customer_model.select()
        self.customer_view = QTableView(); self.customer_view.setModel(self.customer_model)
        self.customer_view.setEditTriggers(QTableView.EditTrigger.DoubleClicked | QTableView.EditTrigger.SelectedClicked)
        self.customer_view.search_text = ""; self.customer_view.setItemDelegate(HighlightDelegate(self.customer_view, self.customer_view))
        self.customer_model.dataChanged.connect(lambda a, b, c=None: self.customer_view.resizeColumnsToContents())
        self.customer_view.resizeColumnsToContents()
        v.addWidget(self.customer_view)
        self.tabs.addTab(tab, "Customers")

    def on_customer_search(self, text):
        t = text.replace("'", "''")
        if t:
            filter_expr = (
                f"first_name LIKE '%{t}%' OR last_name LIKE '%{t}%' OR phone LIKE '%{t}%' "
                f"OR email LIKE '%{t}%' OR address LIKE '%{t}%' OR notes LIKE '%{t}%'"
            )
        else:
            filter_expr = ""
        self.customer_model.setFilter(filter_expr)
        self.customer_view.search_text = text
        self.customer_view.viewport().update()
        self.customer_view.resizeColumnsToContents()

    def new_customer(self):
        conn = sqlite3.connect(DB_FILE); cur = conn.cursor()
        cur.execute("INSERT INTO customers (first_name, last_name) VALUES ('','')"); conn.commit(); conn.close()
        self.customer_model.select(); self.customer_view.resizeColumnsToContents()

    def delete_customer(self):
        sel = self.customer_view.selectionModel().currentIndex()
        if not sel.isValid():
            QMessageBox.information(self, "Select", "Please select a customer row first."); return
        row = sel.row(); idx = self.customer_model.index(row, 0); cid = self.customer_model.data(idx)
        if not cid:
            QMessageBox.warning(self, "Missing", "Selected row has no ID."); return
        if QMessageBox.question(self, "Confirm", f"Delete customer ID {cid} and related records?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) != QMessageBox.StandardButton.Yes:
            return
        conn = sqlite3.connect(DB_FILE); cur = conn.cursor(); cur.execute("DELETE FROM customers WHERE id=?", (cid,)); conn.commit(); conn.close()
        self.customer_model.select(); self.customer_view.resizeColumnsToContents()

    # ---------- Equipment ----------
    def setup_equipment_tab(self):
        tab = QWidget(); v = QVBoxLayout(tab)
        top = QHBoxLayout()
        top.addWidget(QLabel("Search:"))
        self.equipment_search = QLineEdit(); self.equipment_search.setPlaceholderText("serial/cpu/ram/storage/os/notes ...")
        self.equipment_search.textChanged.connect(self.on_equipment_search)
        top.addWidget(self.equipment_search)
        new_btn = QPushButton("New"); del_btn = QPushButton("Delete")
        exp_btn = QPushButton("Export"); imp_btn = QPushButton("Import")
        new_btn.clicked.connect(self.new_equipment); del_btn.clicked.connect(self.delete_equipment)
        exp_btn.clicked.connect(lambda: self._choose_export_and_save("equipment"))
        imp_btn.clicked.connect(lambda: self._import_with_preview("equipment"))
        top.addWidget(new_btn); top.addWidget(del_btn); top.addWidget(exp_btn); top.addWidget(imp_btn)
        v.addLayout(top)
        self.equipment_model = QSqlTableModel(self, self.db); self.equipment_model.setTable("equipment")
        self.equipment_model.setEditStrategy(QSqlTableModel.EditStrategy.OnFieldChange); self.equipment_model.select()
        self.equipment_view = QTableView(); self.equipment_view.setModel(self.equipment_model)
        self.equipment_view.setEditTriggers(QTableView.EditTrigger.DoubleClicked | QTableView.EditTrigger.SelectedClicked)
        self.equipment_view.search_text = ""; self.equipment_view.setItemDelegate(HighlightDelegate(self.equipment_view, self.equipment_view))
        self.equipment_model.dataChanged.connect(lambda a, b, c=None: self.equipment_view.resizeColumnsToContents())
        self.equipment_view.resizeColumnsToContents()
        v.addWidget(self.equipment_view)
        self.tabs.addTab(tab, "Equipment")

    def on_equipment_search(self, text):
        t = text.replace("'", "''")
        if t:
            filter_expr = (
                f"serial_number LIKE '%{t}%' OR cpu LIKE '%{t}%' OR ram LIKE '%{t}%' "
                f"OR storage LIKE '%{t}%' OR os LIKE '%{t}%' OR notes LIKE '%{t}%'"
            )
        else:
            filter_expr = ""
        self.equipment_model.setFilter(filter_expr)
        self.equipment_view.search_text = text
        self.equipment_view.viewport().update()
        self.equipment_view.resizeColumnsToContents()

    def new_equipment(self):
        conn = sqlite3.connect(DB_FILE); cur = conn.cursor(); cur.execute("INSERT INTO equipment (serial_number) VALUES ('')"); conn.commit(); conn.close()
        self.equipment_model.select(); self.equipment_view.resizeColumnsToContents()

    def delete_equipment(self):
        sel = self.equipment_view.selectionModel().currentIndex()
        if not sel.isValid():
            QMessageBox.information(self, "Select", "Please select an equipment row first."); return
        row = sel.row(); idx = self.equipment_model.index(row, 0); eid = self.equipment_model.data(idx)
        if not eid:
            QMessageBox.warning(self, "Missing", "Selected row has no ID."); return
        if QMessageBox.question(self, "Confirm", f"Delete equipment ID {eid}?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) != QMessageBox.StandardButton.Yes:
            return
        conn = sqlite3.connect(DB_FILE); cur = conn.cursor(); cur.execute("DELETE FROM equipment WHERE id=?", (eid,)); conn.commit(); conn.close()
        self.equipment_model.select(); self.equipment_view.resizeColumnsToContents()

    # ---------- Work Orders ----------
    def setup_workorders_tab(self):
        tab = QWidget(); v = QVBoxLayout(tab)
        top = QHBoxLayout()
        top.addWidget(QLabel("Search:"))
        self.wo_search = QLineEdit(); self.wo_search.setPlaceholderText("description/status/... ")
        self.wo_search.textChanged.connect(self.on_workorder_search)
        top.addWidget(self.wo_search)
        new_btn = QPushButton("New"); del_btn = QPushButton("Delete")
        exp_btn = QPushButton("Export"); imp_btn = QPushButton("Import")
        new_btn.clicked.connect(self.new_workorder); del_btn.clicked.connect(self.delete_workorder)
        exp_btn.clicked.connect(lambda: self._choose_export_and_save("work_orders"))
        imp_btn.clicked.connect(lambda: self._import_with_preview("work_orders"))
        top.addWidget(new_btn); top.addWidget(del_btn); top.addWidget(exp_btn); top.addWidget(imp_btn)
        self.wo_pdf_btn = QPushButton("Export WO to PDF"); self.wo_pdf_btn.clicked.connect(self.export_selected_workorder_pdf)
        top.addWidget(self.wo_pdf_btn)
        v.addLayout(top)
        self.wo_model = QSqlTableModel(self, self.db); self.wo_model.setTable("work_orders")
        self.wo_model.setEditStrategy(QSqlTableModel.EditStrategy.OnFieldChange); self.wo_model.select()
        self.wo_view = QTableView(); self.wo_view.setModel(self.wo_model)
        self.wo_view.setEditTriggers(QTableView.EditTrigger.DoubleClicked | QTableView.EditTrigger.SelectedClicked)
        self.wo_view.search_text = ""; self.wo_view.setItemDelegate(HighlightDelegate(self.wo_view, self.wo_view))
        status_col = self.wo_model.fieldIndex("status")
        combo_items = ["Pending", "In Progress", "Completed"]
        color_map = {"pending": QColor(255, 200, 200), "in progress": QColor(255, 255, 200), "completed": QColor(200, 255, 200)}
        class ComboWithColor(ComboBoxDelegate):
            def __init__(self, items, cmap, parent=None):
                super().__init__(items, parent)
                self.cmap = {k.lower(): v for k, v in cmap.items()}
            def initStyleOption(self, option, index):
                super().initStyleOption(option, index)
                val = index.data(Qt.ItemDataRole.DisplayRole)
                if isinstance(val, str):
                    v = val.lower()
                    if v in self.cmap:
                        option.backgroundBrush = QBrush(self.cmap[v])
        combo_delegate = ComboWithColor(combo_items, color_map, self.wo_view)
        self.wo_view.setItemDelegateForColumn(status_col, combo_delegate)
        self.wo_model.dataChanged.connect(lambda a, b, c=None: self.wo_view.resizeColumnsToContents())
        self.wo_view.resizeColumnsToContents()
        v.addWidget(self.wo_view)
        self.tabs.addTab(tab, "Work Orders")

    def on_workorder_search(self, text):
        t = text.replace("'", "''")
        if t:
            filter_expr = f"description LIKE '%{t}%' OR status LIKE '%{t}%'"
        else:
            filter_expr = ""
        self.wo_model.setFilter(filter_expr)
        self.wo_view.search_text = text
        self.wo_view.viewport().update()
        self.wo_view.resizeColumnsToContents()

    def new_workorder(self):
        created = datetime.now().strftime("%Y-%m-%d")
        conn = sqlite3.connect(DB_FILE); cur = conn.cursor()
        cur.execute("INSERT INTO work_orders (description, date_created, status) VALUES ('', ?, 'Pending')", (created,))
        conn.commit(); conn.close()
        self.wo_model.select(); self.wo_view.resizeColumnsToContents()

    def delete_workorder(self):
        sel = self.wo_view.selectionModel().currentIndex()
        if not sel.isValid():
            QMessageBox.information(self, "Select", "Please select a work order row first."); return
        row = sel.row(); idx = self.wo_model.index(row, 0); wid = self.wo_model.data(idx)
        if not wid:
            QMessageBox.warning(self, "Missing", "Selected row has no ID."); return
        if QMessageBox.question(self, "Confirm", f"Delete work order ID {wid} and related invoices?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) != QMessageBox.StandardButton.Yes:
            return
        conn = sqlite3.connect(DB_FILE); cur = conn.cursor(); cur.execute("DELETE FROM work_orders WHERE id=?", (wid,)); conn.commit(); conn.close()
        self.wo_model.select(); self.wo_view.resizeColumnsToContents()

    def export_selected_workorder_pdf(self):
        sel = self.wo_view.selectionModel().currentIndex()
        if not sel.isValid():
            QMessageBox.information(self, "Select", "Please select a work order row first."); return
        row = sel.row(); idx = self.wo_model.index(row, 0); wid = self.wo_model.data(idx)
        if not wid:
            QMessageBox.warning(self, "Missing", "Selected row has no ID."); return
        fname, _ = QFileDialog.getSaveFileName(self, "Save Work Order PDF", f"WorkOrder_{wid}.pdf", "PDF Files (*.pdf)")
        if not fname: return
        if not HAVE_REPORTLAB:
            QMessageBox.critical(self, "Missing dependency", "ReportLab not installed. Run: pip install reportlab"); return
        try:
            generate_workorder_pdf(int(wid), fname)
            QMessageBox.information(self, "Saved", f"Work order PDF saved to {fname}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to generate PDF: {e}")

    # ---------- Invoices ----------
    def setup_invoices_tab(self):
        tab = QWidget(); v = QVBoxLayout(tab)
        top = QHBoxLayout()
        top.addWidget(QLabel("Search:"))
        self.inv_search = QLineEdit(); self.inv_search.setPlaceholderText("status/due_date/amount/notes ...")
        self.inv_search.textChanged.connect(self.on_invoice_search)
        top.addWidget(self.inv_search)
        new_btn = QPushButton("New"); del_btn = QPushButton("Delete")
        exp_btn = QPushButton("Export"); imp_btn = QPushButton("Import")
        new_btn.clicked.connect(self.new_invoice); del_btn.clicked.connect(self.delete_invoice)
        exp_btn.clicked.connect(lambda: self._choose_export_and_save("invoices"))
        imp_btn.clicked.connect(lambda: self._import_with_preview("invoices"))
        top.addWidget(new_btn); top.addWidget(del_btn); top.addWidget(exp_btn); top.addWidget(imp_btn)
        self.inv_pdf_btn = QPushButton("Export Invoice to PDF"); self.inv_pdf_btn.clicked.connect(self.export_selected_invoice_pdf)
        top.addWidget(self.inv_pdf_btn)
        v.addLayout(top)
        self.inv_model = QSqlTableModel(self, self.db); self.inv_model.setTable("invoices")
        self.inv_model.setEditStrategy(QSqlTableModel.EditStrategy.OnFieldChange); self.inv_model.select()
        self.inv_view = QTableView(); self.inv_view.setModel(self.inv_model)
        self.inv_view.setEditTriggers(QTableView.EditTrigger.DoubleClicked | QTableView.EditTrigger.SelectedClicked)
        self.inv_view.search_text = ""; self.inv_view.setItemDelegate(HighlightDelegate(self.inv_view, self.inv_view))
        status_col = self.inv_model.fieldIndex("status")
        inv_items = ["Outstanding", "Paid"]
        inv_color_map = {"outstanding": QColor(255, 200, 200), "paid": QColor(200, 255, 200)}
        class InvComboWithColor(ComboBoxDelegate):
            def __init__(self, items, cmap, parent=None):
                super().__init__(items, parent)
                self.cmap = {k.lower(): v for k, v in cmap.items()}
            def initStyleOption(self, option, index):
                super().initStyleOption(option, index)
                val = index.data(Qt.ItemDataRole.DisplayRole)
                if isinstance(val, str):
                    v = val.lower()
                    if v in self.cmap:
                        option.backgroundBrush = QBrush(self.cmap[v])
        inv_delegate = InvComboWithColor(inv_items, inv_color_map, self.inv_view)
        self.inv_view.setItemDelegateForColumn(status_col, inv_delegate)
        self.inv_model.dataChanged.connect(lambda a, b, c=None: self.inv_view.resizeColumnsToContents())
        self.inv_view.resizeColumnsToContents()
        v.addWidget(self.inv_view)
        self.tabs.addTab(tab, "Invoices")

    def on_invoice_search(self, text):
        t = text.replace("'", "''")
        if t:
            filter_expr = f"status LIKE '%{t}%' OR due_date LIKE '%{t}%' OR CAST(amount AS TEXT) LIKE '%{t}%' OR notes LIKE '%{t}%'"
        else:
            filter_expr = ""
        self.inv_model.setFilter(filter_expr)
        self.inv_view.search_text = text
        self.inv_view.viewport().update()
        self.inv_view.resizeColumnsToContents()

    def new_invoice(self):
        conn = sqlite3.connect(DB_FILE); cur = conn.cursor(); cur.execute("INSERT INTO invoices (amount, status) VALUES (0, 'Outstanding')"); conn.commit(); conn.close()
        self.inv_model.select(); self.inv_view.resizeColumnsToContents()

    def delete_invoice(self):
        sel = self.inv_view.selectionModel().currentIndex()
        if not sel.isValid():
            QMessageBox.information(self, "Select", "Please select an invoice row first."); return
        row = sel.row(); idx = self.inv_model.index(row, 0); iid = self.inv_model.data(idx)
        if not iid:
            QMessageBox.warning(self, "Missing", "Selected invoice has no ID."); return
        if QMessageBox.question(self, "Confirm", f"Delete invoice ID {iid}?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) != QMessageBox.StandardButton.Yes:
            return
        conn = sqlite3.connect(DB_FILE); cur = conn.cursor(); cur.execute("DELETE FROM invoices WHERE id=?", (iid,)); conn.commit(); conn.close()
        self.inv_model.select(); self.inv_view.resizeColumnsToContents()

    def export_selected_invoice_pdf(self):
        sel = self.inv_view.selectionModel().currentIndex()
        if not sel.isValid():
            QMessageBox.information(self, "Select", "Please select an invoice row first."); return
        row = sel.row(); idx = self.inv_model.index(row, 0); iid = self.inv_model.data(idx)
        if not iid:
            QMessageBox.warning(self, "Missing", "Selected invoice has no ID."); return
        fname, _ = QFileDialog.getSaveFileName(self, "Save Invoice PDF", f"Invoice_{iid}.pdf", "PDF Files (*.pdf)")
        if not fname: return
        if not HAVE_REPORTLAB:
            QMessageBox.critical(self, "Missing dependency", "ReportLab not installed. Run: pip install reportlab"); return
        try:
            generate_invoice_pdf(int(iid), fname)
            QMessageBox.information(self, "Saved", f"Invoice PDF saved to {fname}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to generate PDF: {e}")

    # ---------- Business Info ----------
    def setup_businessinfo_tab(self):
        tab = QWidget(); v = QVBoxLayout(tab)
        v.addWidget(QLabel("Business info (used on PDFs)"))
        self.biz_name = QLineEdit()
        self.biz_address = QTextEdit(); self.biz_address.setFixedHeight(100)
        self.biz_phone = QLineEdit(); self.biz_email = QLineEdit(); self.biz_website = QLineEdit()
        self.load_business_info()
        v.addWidget(QLabel("Name:")); v.addWidget(self.biz_name)
        v.addWidget(QLabel("Address:")); v.addWidget(self.biz_address)
        v.addWidget(QLabel("Phone:")); v.addWidget(self.biz_phone)
        v.addWidget(QLabel("Email:")); v.addWidget(self.biz_email)
        v.addWidget(QLabel("Website:")); v.addWidget(self.biz_website)
        btn_layout = QHBoxLayout()
        save_btn = QPushButton("Save"); save_btn.clicked.connect(self.save_business_info)
        reset_btn = QPushButton("Clear"); reset_btn.clicked.connect(self.clear_business_info)
        restore_btn = QPushButton("Restore Defaults"); restore_btn.clicked.connect(self.restore_business_defaults)
        btn_layout.addWidget(save_btn); btn_layout.addWidget(reset_btn); btn_layout.addWidget(restore_btn)
        v.addLayout(btn_layout)
        self.tabs.addTab(tab, "Business Info")

    def load_business_info(self):
        conn = sqlite3.connect(DB_FILE); cur = conn.cursor()
        cur.execute("SELECT name, address, phone, email, website FROM business_info WHERE id=1")
        rec = cur.fetchone() or ("", "", "", "", "")
        conn.close()
        self.biz_name.setText(rec[0] or "")
        self.biz_address.setPlainText(rec[1] or "")
        self.biz_phone.setText(rec[2] or "")
        self.biz_email.setText(rec[3] or "")
        self.biz_website.setText(rec[4] or "")

    def save_business_info(self):
        name = self.biz_name.text().strip(); address = self.biz_address.toPlainText().strip()
        phone = self.biz_phone.text().strip(); email = self.biz_email.text().strip(); website = self.biz_website.text().strip()
        conn = sqlite3.connect(DB_FILE); cur = conn.cursor()
        cur.execute("UPDATE business_info SET name=?, address=?, phone=?, email=?, website=? WHERE id=1", (name, address, phone, email, website))
        conn.commit(); conn.close()
        QMessageBox.information(self, "Saved", "Business info saved.")

    def clear_business_info(self):
        self.biz_name.setText(""); self.biz_address.setPlainText(""); self.biz_phone.setText(""); self.biz_email.setText(""); self.biz_website.setText("")
        self.save_business_info()

    def restore_business_defaults(self):
        default_name = "Business Name Here"; default_addr = "Address line 1\nAddress line 2"
        default_phone = "Phone"; default_email = "email@example.com"; default_web = "https://example.com"
        self.biz_name.setText(default_name); self.biz_address.setPlainText(default_addr)
        self.biz_phone.setText(default_phone); self.biz_email.setText(default_email); self.biz_website.setText(default_web)
        self.save_business_info()

    def closeEvent(self, event):
        try:
            ensure_backups_folder()
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            dst = os.path.join(BACKUP_DIR, f"crm_exit_backup_{ts}.sqlite")
            if os.path.exists(DB_FILE):
                shutil.copy2(DB_FILE, dst)
                print(f"Exit backup created: {dst}")
        except Exception as e:
            print("Exit backup failed:", e)
        super().closeEvent(event)

# -------------------------
# Bootstrap
# -------------------------
def main():
    ensure_backups_folder()
    app = QApplication(sys.argv)
    win = CRMApp()
    win.show()
    if not HAVE_OPENPYXL:
        print("Note: openpyxl not installed. Excel import/export disabled. Install with: pip install openpyxl")
    if not HAVE_REPORTLAB:
        print("Note: reportlab not installed. PDF export disabled. Install with: pip install reportlab")
    if not HAVE_MPL:
        print("Note: matplotlib not installed. Dashboard charts disabled. Install with: pip install matplotlib")
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
